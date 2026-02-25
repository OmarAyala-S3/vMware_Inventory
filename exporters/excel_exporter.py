"""
Exportador Excel con formato profesional - multi-hoja con colores y estilos
"""
import os
from datetime import datetime
from typing import Optional, Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

C_HEADER_BG = "1E3A5F"
C_HEADER_FG = "FFFFFF"
C_ALT_ROW   = "EBF2FF"
C_TITLE_BG  = "0D2B4E"
C_GREEN     = "D4EDDA"
C_RED       = "F8D7DA"
C_YELLOW    = "FFF3CD"

def _apply_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(fill_type="solid", fgColor=C_HEADER_BG)
        cell.font = Font(bold=True, color=C_HEADER_FG, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

def _auto_width(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

def _add_title(ws, title, num_cols, subtitle=""):
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1)
    c.value = title
    c.font = Font(bold=True, color=C_HEADER_FG, size=14, name="Calibri")
    c.fill = PatternFill(fill_type="solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        s = ws.cell(row=2, column=1)
        s.value = subtitle
        s.font = Font(italic=True, color="8EB4E3", size=9, name="Calibri")
        s.fill = PatternFill(fill_type="solid", fgColor=C_HEADER_BG)
        s.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 15

class ExcelExporter:
    def __init__(self, log_callback: Optional[Callable] = None):
        self._log = log_callback or print

    def export(self, output_path, vms=None, hosts=None, datastores=None,
               networks=None, vcenter_name=""):
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        safe = (vcenter_name or "Export").replace(".", "_").replace(":", "_")
        filename = f"Inventario_{safe}_{ts}.xlsx"
        filepath = os.path.join(output_path, filename)
        self._log(f"[INFO] Generando: {filename}")
        gen_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        sheets = {}
        if vms:
            sheets["Maquinas Virtuales"] = pd.DataFrame([v.to_dict() for v in vms])
        if hosts:
            sheets["Hosts ESXi"] = pd.DataFrame([h.to_dict() for h in hosts])
        if datastores:
            sheets["Datastores"] = pd.DataFrame([d.to_dict() for d in datastores])
        if networks:
            sheets["Redes"] = pd.DataFrame([n.to_dict() for n in networks])

        if not sheets:
            self._log("[WARN] Sin datos para exportar.")
            return ""

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=False, startrow=2)

        wb = load_workbook(filepath)

        if "Maquinas Virtuales" in wb.sheetnames and vms:
            self._fmt_vms(wb["Maquinas Virtuales"], len(vms), gen_at)
        if "Hosts ESXi" in wb.sheetnames and hosts:
            self._fmt_hosts(wb["Hosts ESXi"], len(hosts), gen_at)
        if "Datastores" in wb.sheetnames and datastores:
            self._fmt_ds(wb["Datastores"], len(datastores), gen_at)
        if "Redes" in wb.sheetnames and networks:
            self._fmt_nets(wb["Redes"], len(networks), gen_at)

        self._add_summary(wb, vms, hosts, datastores, networks, vcenter_name, gen_at)

        # Ordenar hojas
        order = ["Resumen", "Maquinas Virtuales", "Hosts ESXi", "Datastores", "Redes"]
        for target_i, name in enumerate(order):
            if name in wb.sheetnames:
                current_i = wb.sheetnames.index(name)
                wb.move_sheet(name, offset=target_i - current_i)

        wb.save(filepath)
        self._log(f"[OK] Archivo guardado: {filepath}")
        return filepath

    def _color_rows(self, ws, header_row, estado_col=None, tools_col=None, pct_col=None):
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            r = row[0].row
            alt = C_ALT_ROW if (r - header_row) % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(vertical="center")
                if estado_col and cell.column == estado_col:
                    v = str(cell.value or "")
                    if v in ("Encendida", "Conectado"):
                        cell.fill = PatternFill(fill_type="solid", fgColor=C_GREEN)
                        cell.font = Font(name="Calibri", size=9, bold=True, color="155724")
                    elif v in ("Apagada", "Desconectado"):
                        cell.fill = PatternFill(fill_type="solid", fgColor=C_RED)
                        cell.font = Font(name="Calibri", size=9, bold=True, color="721C24")
                    elif v == "Suspendida":
                        cell.fill = PatternFill(fill_type="solid", fgColor=C_YELLOW)
                    else:
                        cell.fill = PatternFill(fill_type="solid", fgColor=alt)
                elif tools_col and cell.column == tools_col:
                    v = str(cell.value or "").lower()
                    if "ok" in v:
                        cell.fill = PatternFill(fill_type="solid", fgColor=C_GREEN)
                    elif "old" in v or "outdated" in v:
                        cell.fill = PatternFill(fill_type="solid", fgColor=C_YELLOW)
                    elif "not" in v or "none" in v:
                        cell.fill = PatternFill(fill_type="solid", fgColor=C_RED)
                    else:
                        cell.fill = PatternFill(fill_type="solid", fgColor=alt)
                elif pct_col and cell.column == pct_col:
                    try:
                        pct = float(cell.value or 0)
                        if pct >= 85:
                            cell.fill = PatternFill(fill_type="solid", fgColor=C_RED)
                            cell.font = Font(name="Calibri", size=9, bold=True, color="721C24")
                        elif pct >= 70:
                            cell.fill = PatternFill(fill_type="solid", fgColor=C_YELLOW)
                        else:
                            cell.fill = PatternFill(fill_type="solid", fgColor=C_GREEN)
                    except Exception:
                        cell.fill = PatternFill(fill_type="solid", fgColor=alt)
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor=alt)

    def _fmt_vms(self, ws, count, gen_at):
        nc = ws.max_column
        _add_title(ws, "Inventario de Maquinas Virtuales", nc,
                   f"Generado: {gen_at} | Total VMs: {count}")
        hr = 5
        _apply_header(ws, hr, nc)
        ec = tc = None
        for cell in ws[hr]:
            if cell.value == "Estado":
                ec = cell.column
            if cell.value == "VMware Tools Status":
                tc = cell.column
        self._color_rows(ws, hr, estado_col=ec, tools_col=tc)
        _auto_width(ws)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _fmt_hosts(self, ws, count, gen_at):
        nc = ws.max_column
        _add_title(ws, "Inventario de Hosts ESXi", nc,
                   f"Generado: {gen_at} | Total Hosts: {count}")
        hr = 5
        _apply_header(ws, hr, nc)
        ec = None
        for cell in ws[hr]:
            if cell.value == "Estado":
                ec = cell.column
        self._color_rows(ws, hr, estado_col=ec)
        _auto_width(ws)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _fmt_ds(self, ws, count, gen_at):
        nc = ws.max_column
        _add_title(ws, "Inventario de Datastores", nc,
                   f"Generado: {gen_at} | Total Datastores: {count}")
        hr = 5
        _apply_header(ws, hr, nc)
        pc = None
        for cell in ws[hr]:
            if cell.value == "% Usado":
                pc = cell.column
        self._color_rows(ws, hr, pct_col=pc)
        _auto_width(ws)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _fmt_nets(self, ws, count, gen_at):
        nc = ws.max_column
        _add_title(ws, "Inventario de Redes", nc,
                   f"Generado: {gen_at} | Total Redes: {count}")
        hr = 5
        _apply_header(ws, hr, nc)
        self._color_rows(ws, hr)
        _auto_width(ws)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _add_summary(self, wb, vms, hosts, datastores, networks, vcenter_name, gen_at):
        ws = wb.create_sheet("Resumen")
        tf = PatternFill(fill_type="solid", fgColor=C_TITLE_BG)
        hf = PatternFill(fill_type="solid", fgColor=C_HEADER_BG)
        cf = PatternFill(fill_type="solid", fgColor=C_ALT_ROW)
        gf = PatternFill(fill_type="solid", fgColor=C_GREEN)
        rf = PatternFill(fill_type="solid", fgColor=C_RED)

        def sc(r, c, val, bold=False, sz=10, clr="000000", fill=None, al="left"):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.font = Font(name="Calibri", bold=bold, size=sz, color=clr)
            cell.alignment = Alignment(horizontal=al, vertical="center")
            if fill:
                cell.fill = fill
            return cell

        ws.merge_cells("A1:H1")
        sc(1, 1, "RESUMEN DE INVENTARIO VMWARE", bold=True, sz=16,
           clr=C_HEADER_FG, fill=tf, al="center")
        ws.row_dimensions[1].height = 38

        ws.merge_cells("A2:H2")
        sc(2, 1, f"Fuente: {vcenter_name or 'vCenter/ESXi'}  |  Generado: {gen_at}",
           sz=9, clr="8EB4E3", fill=hf, al="center")
        ws.row_dimensions[2].height = 16

        row = 4
        ws.merge_cells(f"A{row}:H{row}")
        sc(row, 1, "TOTALES", bold=True, sz=11, clr=C_HEADER_FG, fill=hf, al="center")
        ws.row_dimensions[row].height = 22
        row += 1

        def mrow(r, label, value, note="", vfill=None):
            ws.row_dimensions[r].height = 18
            lc = ws.cell(row=r, column=1)
            lc.value = label
            lc.font = Font(name="Calibri", bold=True, size=10)
            lc.fill = cf
            lc.alignment = Alignment(vertical="center")
            ws.merge_cells(f"A{r}:C{r}")
            vc = ws.cell(row=r, column=4)
            vc.value = value
            vc.font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.fill = vfill or PatternFill(fill_type="solid", fgColor="DDEEFF")
            if note:
                nc = ws.cell(row=r, column=5)
                nc.value = note
                nc.font = Font(name="Calibri", italic=True, size=9, color="666666")
                nc.alignment = Alignment(horizontal="left", vertical="center")
                ws.merge_cells(f"E{r}:H{r}")

        mrow(row, "Total Maquinas Virtuales", len(vms) if vms else 0)
        row += 1
        if vms:
            on = sum(1 for v in vms if v.power_state == "Encendida")
            off = sum(1 for v in vms if v.power_state == "Apagada")
            mrow(row, "  Encendidas", on, vfill=gf)
            row += 1
            mrow(row, "  Apagadas", off, vfill=rf)
            row += 1
            mrow(row, "  Otras / Suspendidas", len(vms) - on - off)
            row += 1
        mrow(row, "Total Hosts ESXi", len(hosts) if hosts else 0)
        row += 1
        mrow(row, "Total Datastores", len(datastores) if datastores else 0)
        row += 1
        mrow(row, "Total Redes", len(networks) if networks else 0)
        row += 1

        if vms:
            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            sc(row, 1, "RECURSOS TOTALES CONFIGURADOS (VMs)", bold=True, sz=11,
               clr=C_HEADER_FG, fill=hf, al="center")
            ws.row_dimensions[row].height = 22
            row += 1
            mrow(row, "Total vCPUs", sum(v.vcpu for v in vms))
            row += 1
            total_ram = sum(v.ram_mb for v in vms)
            mrow(row, "Total RAM", f"{round(total_ram/1024,1)} GB",
                 f"= {round(total_ram/1024/1024,2)} TB")
            row += 1
            total_disk = sum(sum(d.size_gb for d in v.disks) for v in vms)
            mrow(row, "Total Storage VM", f"{round(total_disk,1)} GB",
                 f"= {round(total_disk/1024,2)} TB")
            row += 1

        if datastores:
            critical = [d for d in datastores
                        if d.capacity_gb > 0 and (d.used_gb / d.capacity_gb) >= 0.85]
            if critical:
                row += 1
                ws.merge_cells(f"A{row}:H{row}")
                sc(row, 1, f"DATASTORES CRITICOS (>85% uso): {len(critical)}",
                   bold=True, sz=11, clr="721C24", fill=rf, al="center")
                ws.row_dimensions[row].height = 22
                row += 1
                for ds in critical:
                    pct = round(ds.used_gb / ds.capacity_gb * 100, 1)
                    mrow(row, f"  {ds.name}", f"{pct}%",
                         f"{ds.free_gb:.1f} GB libres / {ds.capacity_gb:.1f} GB total",
                         vfill=rf)
                    row += 1

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 4
        ws.column_dimensions["C"].width = 4
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 38
        for col in ["F", "G", "H"]:
            ws.column_dimensions[col].width = 10
