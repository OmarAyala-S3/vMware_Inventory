"""
exporters/multi_exporter.py
Exportador Excel multi-fuente:
  - Una hoja por cada vCenter/Host escaneado
  - Hojas consolidadas al final (todas las VMs, todos los Hosts, etc.)
  - Formato profesional con colores por estado
"""
import os
import logging
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from services.connection_manager import ConsolidatedResult
from models.connection_profile import ConnectionProfile

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────
# Paleta de colores
# ─────────────────────────────────────────────────────────
class Colors:
    HEADER_SOURCE  = "1F4E79"   # Azul oscuro — headers de hoja por fuente
    HEADER_CONSOL  = "833C00"   # Naranja oscuro — headers consolidados
    HEADER_SUMMARY = "375623"   # Verde oscuro — header resumen
    VM_ON          = "E2EFDA"   # Verde claro — VM encendida
    VM_OFF         = "FCE4D6"   # Rojo claro — VM apagada
    VM_SUSPENDED   = "FFF2CC"   # Amarillo — suspendida
    HOST_OK        = "DEEAF1"   # Azul claro — host conectado
    HOST_ERROR     = "FCE4D6"   # Rojo claro — host desconectado
    WHITE          = "FFFFFF"
    LIGHT_GRAY     = "F2F2F2"
    SOURCE_TAB_OK  = "70AD47"   # Verde — tab exitosa
    SOURCE_TAB_ERR = "FF0000"   # Rojo — tab con error

# ─────────────────────────────────────────────────────────
# Mapeo de columnas
# ─────────────────────────────────────────────────────────
VM_COLUMNS = [
    ("Fuente",              "source_name"),
    ("vCenter",             "vcenter"),
    ("Host Físico",         "host"),
    ("Ambiente",            "environment"),
    ("Hostname",            "hostname"),
    ("Descripción",         "description"),
    ("Estado",              "power_state"),
    ("IP",                  "ip_address"),
    ("MAC",                 "mac_address"),
    ("Procesador",          "processor"),
    ("vCPU",                "vcpu"),
    ("RAM (GB)",            "_ram_gb"),          # calculado en _vm_to_row
    ("Discos",              "_disks_str"),        # calculado en _vm_to_row
    ("Storage Total (GB)",  "_storage_gb"),       # calculado en _vm_to_row
    ("Datastore",           "datastore"),
    ("Red",                 "network"),
    ("Dominio",             "domain"),
    ("Sistema Operativo",   "os_name"),
    ("Edición SO",          "os_edition"),
    ("VMware Tools",        "tools_status"),
    ("Versión Tools",       "tools_version"),
    ("Versión HW",          "hw_version"),
]

HOST_COLUMNS = [
    ("Fuente",              "source_name"),
    ("vCenter",             "vcenter"),
    ("Nombre Host",         "name"),
    ("IP",                  "ip_address"),
    ("Versión ESXi",        "esxi_version"),
    ("Build",               "build"),
    ("Fabricante",          "vendor"),
    ("Modelo",              "model"),
    ("CPU Model",           "cpu_model"),
    ("CPU Cores",           "cpu_cores"),
    ("CPU Threads",         "cpu_threads"),
    ("RAM Total (GB)",      "ram_total_gb"),
    ("RAM Usada (GB)",      "ram_used_gb"),
    ("RAM Libre (GB)",      "_ram_libre_gb"),     # calculado en _host_to_row
    ("Estado",              "state"),
    ("Cluster",             "cluster"),
    ("Nº Serie",            "serial_number"),
    ("Datastores",          "_datastores_str"),   # calculado en _host_to_row
]

DATASTORE_COLUMNS = [
    ("Fuente",              "source_name"),
    ("Nombre",              "name"),
    ("Tipo",                "ds_type"),
    ("Capacidad (GB)",      "capacity_gb"),
    ("Libre (GB)",          "free_gb"),
    ("Usado (GB)",          "used_gb"),
    ("% Usado",             "_used_pct"),         # calculado en _ds_to_row
    ("Accesible",           "accessible"),
]

NETWORK_COLUMNS = [
    ("Fuente",              "source_name"),
    ("Nombre",              "name"),
    ("Tipo",                "net_type"),
    ("VLAN",                "vlan_id"),
    ("Switch",              "switch_name"),
    ("VMs Conectadas",      "vms_count"),
    ("Hosts",               "_hosts_str"),        # calculado en _net_to_row
]

class MultiSourceExporter:
    """
    Exporta un ConsolidatedResult a un archivo Excel con:
      - 1 hoja de Resumen ejecutivo
      - N hojas (una por fuente) con secciones VMs + Hosts + Datastores
      - 4 hojas consolidadas al final (todas las VMs, Hosts, Datastores, Redes)
    """

    def __init__(self, output_dir: str = "."):
        self.output_dir = output_dir
        self._wb = None

    def export(
        self,
        consolidated: ConsolidatedResult,
        profiles: List[ConnectionProfile],
        filename: Optional[str] = None,
    ) -> str:
        """
        Genera el archivo Excel y retorna la ruta del archivo creado.
        """
        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            sources = len(consolidated.results_by_source)
            filename = f"Inventario_VMware_{sources}fuentes_{ts}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        # Construir DataFrames por fuente
        source_dfs = self._build_source_dataframes(consolidated, profiles)

        # Construir DataFrames consolidados
        all_vms_df        = self._concat_category(source_dfs, "vms")
        all_hosts_df      = self._concat_category(source_dfs, "hosts")
        all_datastores_df = self._concat_category(source_dfs, "datastores")
        all_networks_df   = self._concat_category(source_dfs, "networks")

        # Escribir con pandas primero (estructura base)
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

            # ── Hoja resumen
            summary_df = self._build_summary_df(consolidated, profiles)
            summary_df.to_excel(writer, sheet_name="📊 Resumen", index=False)

            # ── Una hoja por fuente
            for profile in profiles:
                if profile.id not in source_dfs:
                    continue
                sheet_name = self._safe_sheet_name(profile.alias or profile.host)
                dfs = source_dfs[profile.id]
                self._write_source_sheet(writer, sheet_name, dfs, profile)

            # ── Hojas consolidadas
            if not all_vms_df.empty:
                all_vms_df.to_excel(writer, sheet_name="🖥 Todas las VMs", index=False)
            if not all_hosts_df.empty:
                all_hosts_df.to_excel(writer, sheet_name="⚙ Todos los Hosts", index=False)
            if not all_datastores_df.empty:
                all_datastores_df.to_excel(writer, sheet_name="💾 Datastores", index=False)
            if not all_networks_df.empty:
                all_networks_df.to_excel(writer, sheet_name="🌐 Redes", index=False)

        # ── Post-proceso: aplicar estilos con openpyxl
        self._apply_styles(filepath, consolidated, profiles)

        logger.info(f"Excel exportado: {filepath}")
        return filepath

    # ─────────────────────────────────────────────
    # Construcción de DataFrames
    # ─────────────────────────────────────────────

    def _build_source_dataframes(self, consolidated, profiles) -> Dict:
        """Construye un dict { profile_id: { 'vms': df, 'hosts': df, ... } }"""
        result = {}
        for profile in profiles:
            inv = consolidated.results_by_source.get(profile.id)
            if inv is None:
                continue

            vms_data  = [self._vm_to_row(vm)   for vm  in getattr(inv, 'virtual_machines', [])]
            host_data = [self._host_to_row(h)   for h   in getattr(inv, 'hosts',            [])]
            ds_data   = [self._ds_to_row(ds)    for ds  in getattr(inv, 'datastores',        [])]
            net_data  = [self._net_to_row(n)    for n   in getattr(inv, 'networks',          [])]

            result[profile.id] = {
                "vms":        self._make_df(vms_data,  VM_COLUMNS),
                "hosts":      self._make_df(host_data, HOST_COLUMNS),
                "datastores": self._make_df(ds_data,   DATASTORE_COLUMNS),
                "networks":   self._make_df(net_data,  NETWORK_COLUMNS),
                "profile":    profile,
            }
        return result

    def _make_df(self, rows: list, columns: list) -> pd.DataFrame:
        if rows:
            headers = list(rows[0].keys())
            return pd.DataFrame(rows, columns=headers)
        headers = [col[0] for col in columns if not col[1].startswith("_")]
        return pd.DataFrame(columns=headers)

    def _concat_category(self, source_dfs: dict, category: str) -> pd.DataFrame:
        frames = [
            v[category] for v in source_dfs.values()
            if category in v and not v[category].empty
        ]
        if not frames:
            return pd.DataFrame()
        return pd.concat(frames, ignore_index=True)

    def _build_summary_df(self, consolidated, profiles) -> pd.DataFrame:
        rows = []
        for profile in profiles:
            inv = consolidated.results_by_source.get(profile.id)
            status = "✅ OK" if inv else "❌ Error"
            error  = profile.error_message if profile.has_error else ""
            rows.append({
                "Fuente":       profile.display_name,
                "IP/FQDN":      profile.host,
                "Tipo":         profile.connection_type.value,
                "Estado":       status,
                "VMs":          profile.vms_found,
                "Hosts":        profile.hosts_found,
                "Datastores":   profile.datastores_found,
                "Error":        error,
            })
        # Totales
        rows.append({
            "Fuente":     "TOTAL",
            "IP/FQDN":    "",
            "Tipo":       "",
            "Estado":     f"{len(consolidated.completed_profiles)} exitosas / {len(consolidated.failed_profiles)} fallidas",
            "VMs":        consolidated.total_vms,
            "Hosts":      consolidated.total_hosts,
            "Datastores": consolidated.total_datastores,
            "Error":      "",
        })
        return pd.DataFrame(rows)

    def _write_source_sheet(self, writer, sheet_name, dfs, profile):
        """
        Escribe una hoja por fuente con secciones separadas:
        VMs → espacio → Hosts → espacio → Datastores → Redes
        """
        workbook  = writer.book
        worksheet = workbook.create_sheet(title=sheet_name)

        row_cursor = 1

        sections = [
            ("🖥 Máquinas Virtuales",    dfs["vms"]),
            ("⚙ Hosts ESXi",             dfs["hosts"]),
            ("💾 Datastores",            dfs["datastores"]),
            ("🌐 Redes",                 dfs["networks"]),
        ]

        for section_title, df in sections:
            if df.empty:
                continue

            # Título de sección
            worksheet.cell(row=row_cursor, column=1, value=section_title)
            row_cursor += 1

            # Headers
            for col_idx, col_name in enumerate(df.columns, start=1):
                worksheet.cell(row=row_cursor, column=col_idx, value=col_name)
            row_cursor += 1

            # Datos
            for _, data_row in df.iterrows():
                for col_idx, value in enumerate(data_row, start=1):
                    worksheet.cell(row=row_cursor, column=col_idx, value=value)
                row_cursor += 1

            row_cursor += 2  # Espacio entre secciones

    # ─────────────────────────────────────────────
    # Conversores objeto → dict
    # ─────────────────────────────────────────────

    def _vm_to_row(self, vm) -> dict:
        row = {}

        # ── Campos directos del modelo ────────────────────────────────────
        direct_fields = [
            ("Fuente",           "source_name"),
            ("vCenter",          "vcenter"),
            ("Host Físico",      "host"),
            ("Ambiente",         "environment"),
            ("Hostname",         "hostname"),
            ("Descripción",      "description"),
            ("Estado",           "power_state"),
            ("Procesador",       "processor"),
            ("vCPU",             "vcpu"),
            ("Datastore",        "datastore"),
            ("Red",              "network"),
            ("Dominio",          "domain"),
            ("Sistema Operativo","os_name"),
            ("Edición SO",       "os_edition"),
            ("VMware Tools",     "tools_status"),
            ("Versión Tools",    "tools_version"),
            ("Versión HW",       "hw_version"),
        ]
        for label, attr in direct_fields:
            row[label] = getattr(vm, attr, "") or ""

        # ── IP: primero desde NICs, luego campo directo ───────────────────
        nics = getattr(vm, 'nics', []) or []
        first_ip  = ""
        first_mac = ""
        first_net = ""
        for nic in nics:
            ips = getattr(nic, 'ip_addresses', []) or []
            if not first_ip and ips:
                first_ip = ips[0]
            if not first_mac:
                first_mac = getattr(nic, 'mac_address', "") or ""
            if not first_net:
                first_net = getattr(nic, 'network', "") or ""
            if first_ip and first_mac and first_net:
                break

        row["IP"]  = first_ip  or getattr(vm, 'ip_address',  "") or ""
        row["MAC"] = first_mac or getattr(vm, 'mac_address', "") or ""
        if not row["Red"] and first_net:
            row["Red"] = first_net

        # ── RAM: ram_mb tiene prioridad, fallback a ram_gb ────────────────
        ram_mb = getattr(vm, 'ram_mb', 0) or 0
        ram_gb = getattr(vm, 'ram_gb', 0.0) or 0.0
        row["RAM (GB)"] = round(ram_mb / 1024, 2) if ram_mb > 0 else round(ram_gb, 2)

        # ── Discos ────────────────────────────────────────────────────────
        disks = getattr(vm, 'disks', []) or []
        if disks:
            row["Discos"] = " | ".join(
                f"{getattr(d,'label','Disk')}: {getattr(d,'size_gb',0):.0f}GB"
                for d in disks
            )
            row["Storage Total (GB)"] = round(
                sum(getattr(d, 'size_gb', 0) for d in disks), 2
            )
        else:
            row["Discos"] = ""
            row["Storage Total (GB)"] = 0.0

        return row

    def _host_to_row(self, host) -> dict:
        row = {}

        direct_fields = [
            ("Fuente",         "source_name"),
            ("vCenter",        "vcenter"),
            ("Nombre Host",    "name"),
            ("IP",             "ip_address"),
            ("Versión ESXi",   "esxi_version"),
            ("Build",          "build"),
            ("Fabricante",     "vendor"),
            ("Modelo",         "model"),
            ("CPU Model",      "cpu_model"),
            ("CPU Cores",      "cpu_cores"),
            ("CPU Threads",    "cpu_threads"),
            ("Estado",         "state"),
            ("Cluster",        "cluster"),
            ("Nº Serie",       "serial_number"),
        ]
        for label, attr in direct_fields:
            row[label] = getattr(host, attr, "") or ""

        # ── RAM ───────────────────────────────────────────────────────────
        ram_total = getattr(host, 'ram_total_gb', 0.0) or 0.0
        ram_used  = getattr(host, 'ram_used_gb',  0.0) or 0.0
        row["RAM Total (GB)"] = round(ram_total, 2)
        row["RAM Usada (GB)"] = round(ram_used,  2)
        row["RAM Libre (GB)"] = round(ram_total - ram_used, 2)

        # ── Datastores ────────────────────────────────────────────────────
        ds_list = getattr(host, 'datastores', []) or []
        row["Datastores"] = " | ".join(ds_list) if ds_list else ""

        return row

    def _ds_to_row(self, ds) -> dict:
        row = {}

        direct_fields = [
            ("Fuente",   "source_name"),
            ("Nombre",   "name"),
            ("Tipo",     "ds_type"),
            ("Accesible","accessible"),
        ]
        for label, attr in direct_fields:
            val = getattr(ds, attr, "") 
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            row[label] = val or ""

        cap  = getattr(ds, 'capacity_gb', 0.0) or 0.0
        free = getattr(ds, 'free_gb',     0.0) or 0.0
        used = getattr(ds, 'used_gb',     0.0) or 0.0

        # Si used_gb no viene calculado, derivarlo
        if used == 0.0 and cap > 0:
            used = cap - free

        row["Capacidad (GB)"] = round(cap,  2)
        row["Libre (GB)"]     = round(free, 2)
        row["Usado (GB)"]     = round(used, 2)
        row["% Usado"]        = f"{(used/cap*100):.1f}%" if cap > 0 else "N/A"

        return row

    def _net_to_row(self, net) -> dict:
        row = {}

        direct_fields = [
            ("Fuente",        "source_name"),
            ("Nombre",        "name"),
            ("Tipo",          "net_type"),
            ("VLAN",          "vlan_id"),
            ("Switch",        "switch_name"),
            ("VMs Conectadas","vms_count"),
        ]
        for label, attr in direct_fields:
            row[label] = getattr(net, attr, "") or ""

        hosts = getattr(net, 'hosts', []) or []
        row["Hosts"] = " | ".join(hosts) if hosts else ""

        return row

    # ─────────────────────────────────────────────
    # Estilos openpyxl
    # ─────────────────────────────────────────────

    def _apply_styles(self, filepath: str, consolidated, profiles):
        wb = load_workbook(filepath)

        for ws in wb.worksheets:
            name = ws.title

            if "Resumen" in name:
                self._style_summary_sheet(ws)
            elif "Todas las VMs" in name:
                self._style_consolidated_sheet(ws, Colors.HEADER_CONSOL, "vm")
            elif "Todos los Hosts" in name:
                self._style_consolidated_sheet(ws, Colors.HEADER_CONSOL, "host")
            elif "Datastores" in name or "Redes" in name:
                self._style_consolidated_sheet(ws, Colors.HEADER_CONSOL, "plain")
            else:
                # Hoja por fuente
                self._style_source_sheet(ws)

            # Ajuste de columnas en todos
            self._auto_fit_columns(ws)

        wb.save(filepath)

    def _style_summary_sheet(self, ws):
        header_fill = PatternFill("solid", fgColor=Colors.HEADER_SUMMARY)
        header_font = Font(color="FFFFFF", bold=True, size=11)
        total_fill  = PatternFill("solid", fgColor="D9E1F2")
        total_font  = Font(bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Fila de totales (última)
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.fill = total_fill
            cell.font = total_font

        # Color por estado
        for row in ws.iter_rows(min_row=2, max_row=last_row - 1):
            estado_cell = row[3]  # columna "Estado"
            val = str(estado_cell.value or "")
            if "✅" in val:
                for c in row:
                    c.fill = PatternFill("solid", fgColor="E2EFDA")
            elif "❌" in val:
                for c in row:
                    c.fill = PatternFill("solid", fgColor="FCE4D6")

        self._add_borders(ws)

    def _style_consolidated_sheet(self, ws, header_color: str, category: str):
        header_fill = PatternFill("solid", fgColor=header_color)
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        if category == "vm":
            self._color_vm_rows(ws)
        elif category == "host":
            self._color_host_rows(ws)

        self._add_borders(ws)
        self._freeze_panes(ws, "B2")

    def _style_source_sheet(self, ws):
        """Estiliza hoja multi-sección de una fuente."""
        section_fill = PatternFill("solid", fgColor=Colors.HEADER_SOURCE)
        section_font = Font(color="FFFFFF", bold=True, size=12)
        header_fill  = PatternFill("solid", fgColor="BDD7EE")
        header_font  = Font(bold=True)

        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "")
                # Detectar títulos de sección (emojis)
                if any(e in val for e in ["🖥", "⚙", "💾", "🌐"]):
                    cell.fill = section_fill
                    cell.font = section_font
                    ws.row_dimensions[cell.row].height = 22
                # Detectar headers (coinciden con nombres de columnas conocidos)
                elif val in {col[0] for col in VM_COLUMNS + HOST_COLUMNS + DATASTORE_COLUMNS + NETWORK_COLUMNS}:
                    cell.fill = header_fill
                    cell.font = header_font

        self._add_borders(ws)

    def _color_vm_rows(self, ws):
        on_fill   = PatternFill("solid", fgColor=Colors.VM_ON)
        off_fill  = PatternFill("solid", fgColor=Colors.VM_OFF)
        sus_fill  = PatternFill("solid", fgColor=Colors.VM_SUSPENDED)

        # Buscar columna "Estado"
        estado_col = None
        for cell in ws[1]:
            if cell.value == "Estado":
                estado_col = cell.column
                break

        if not estado_col:
            return

        for row in ws.iter_rows(min_row=2):
            estado = str(row[estado_col - 1].value or "").lower()
            if "on" in estado or "encendid" in estado or "poweredon" in estado:
                fill = on_fill
            elif "suspend" in estado:
                fill = sus_fill
            else:
                fill = off_fill
            for cell in row:
                cell.fill = fill

    def _color_host_rows(self, ws):
        ok_fill  = PatternFill("solid", fgColor=Colors.HOST_OK)
        err_fill = PatternFill("solid", fgColor=Colors.HOST_ERROR)

        estado_col = None
        for cell in ws[1]:
            if cell.value == "Estado":
                estado_col = cell.column
                break

        if not estado_col:
            return

        for row in ws.iter_rows(min_row=2):
            estado = str(row[estado_col - 1].value or "").lower()
            fill = ok_fill if "connect" in estado else err_fill
            for cell in row:
                cell.fill = fill

    def _add_borders(self, ws):
        thin = Side(border_style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    def _freeze_panes(self, ws, cell: str):
        ws.freeze_panes = cell

    def _auto_fit_columns(self, ws, max_width: int = 50):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception as _e:
                    logger.debug("Excepcion ignorada: %s", _e)
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)

    # ─────────────────────────────────────────────
    # Utilidades
    # ─────────────────────────────────────────────

    def _safe_sheet_name(self, name: str) -> str:
        """Limpia caracteres inválidos para nombres de hoja Excel."""
        invalid = r'\/*?:[]'
        for ch in invalid:
            name = name.replace(ch, "_")
        return name[:31]  # Límite de Excel
